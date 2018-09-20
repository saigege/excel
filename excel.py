# -*- coding: utf-8 -*-
"""
python3.6
Windows
xlrd 1.1.0
openpyxl 2.5.6
"""
import xlrd
import os
import openpyxl

#路径和文件名
excelPath = os.path.join(os.getcwd(), 'comepelet.xlsx') 
workbook = xlrd.open_workbook(r'D:\PythonWorkSpace\TTT\InputData.xlsx')
sheet2 = workbook.sheet_by_index(1) # sheet索引从0开始
sheet1 = workbook.sheet_by_name('Sheet1')
Pucks = workbook.sheet_by_name('Pucks')

x = openpyxl.Workbook()
sheet2 = x.active
sheet2.title = 'test'
   
#获取列数    
nrows1 = sheet1.nrows
nrows2 = Pucks.nrows
for i in range(1,nrows1):#旅客遍历    
    for j in range(1,nrows2):#航班号遍历 
        sheet2.cell(row=i, column=1, value=sheet1.cell(i,0).value)#旅客记录号
        sheet2.cell(row=i, column=2, value=sheet1.cell(i,1).value)#乘客数
        sheet2.cell(row=i, column=3, value=sheet1.cell(i,2).value)#到达航班
        sheet2.cell(row=i, column=8, value=sheet1.cell(i,8).value)#出发航班
        #航班号对应，且日期为20
        if (sheet1.cell(i,2).value == Pucks.cell(j,3).value) and (xlrd.xldate_as_tuple(Pucks.cell_value(j,1),workbook.datemode)[2]==20):
            ###########处理日期##########
            ctype = Pucks.cell(j, 2).ctype
            if ctype == 3:
                time1=xlrd.xldate_as_tuple(Pucks.cell(j,2).value, 1)
                time1=str(time1[3])+':'+str(time1[4])
            else:
                time1 = Pucks.cell(j,2).value
            sheet2.cell(row=i, column=4, value=time1)#到达时刻
            sheet2.cell(row=i, column=5, value=str(Pucks.cell(j,4).value))#到达类型
            sheet2.cell(row=i, column=6, value=str(Pucks.cell(j,5).value))#飞机型号
            
            #判断机体类型
            l1=str(Pucks.cell(j,5).value)
            if (l1=='332'):
                body1='Wide-body'
            elif (l1=='333'):
                body1='Wide-body'
            elif (l1=='33E'):
                body1='Wide-body'
            elif (l1=='33H'):
                body1='Wide-body'
            elif (l1=='33L'):
                body1='Wide-body'
            elif (l1=='773'):
                body1='Wide-body'
            else:
                body1='Narrow-body'
            sheet2.cell(row=i, column=7, value=str(body1))#机体类型
            
    for k in range(1,nrows2):
        date_value2 = xlrd.xldate_as_tuple(Pucks.cell_value(k,6),workbook.datemode)
        date_value2=date_value2[2]
        if sheet1.cell(i,8).value == Pucks.cell(k,8).value  and (date_value2 == 20):
            ctype = Pucks.cell(j, 7).ctype
            if ctype == 3:
                time1=xlrd.xldate_as_tuple(Pucks.cell(j,7).value, 1)
                time1=str(time1[3])+':'+str(time1[4])
            else:
                time1 = Pucks.cell(j,7).value
            sheet2.cell(row=i, column=11, value=time1)#出发时刻
            #sheet2.cell(row=i, column=9, value=Pucks.cell(j,7).value)#出发时刻
            sheet2.cell(row=i, column=9, value=str(Pucks.cell(k,9).value))#出发类型
            sheet2.cell(row=i, column=10, value=str(Pucks.cell(k,5).value))#飞机型号
            
            b2=str(Pucks.cell(k,9).value)
            l2=str(Pucks.cell(k,5).value)
            if (l2=='332'):
                body2='Wide-body'
            elif (l2=='333'):
                body2='Wide-body'
            elif (l2=='33E'):
                body2='Wide-body'
            elif (l2=='33H'):
                body2='Wide-body'
            elif (l2=='33L'):
                body2='Wide-body'
            elif (l2=='773'):
                body2='Wide-body'
            else:
                body2='Narrow-body'
            sheet2.cell(row=i, column=12, value=str(body2))#机体型号
            
x.save(excelPath)   
